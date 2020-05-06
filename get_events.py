#!/usr/bin/python3
# -*- coding: UTF-8 -*-


from pyzabbix import ZabbixAPI
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import date, datetime, timedelta
from pyfiglet import Figlet
from progress.bar import FillingSquaresBar
import argparse, time


# Coleta de parametros
parser = argparse.ArgumentParser(description="Extracao de eventos do Zabbix")
parser.add_argument("-u", "--user", required=True, help="Usuario do Zabbix (Admin)")
parser.add_argument("-p", "--password", required=True, help="Senha do Zabbix (zabbix)")
parser.add_argument(
    "-s", "--server", required=True, help="Endereco Zabbix (http://localhost/zabbix)"
)
parser.add_argument(
    "-n",
    "--name",
    default="Zabbix",
    help="Nome do relatorio gerado ex: Alertas_DATAHOJE.xlsx",
)
parser.add_argument("-g", "--group", action="append", help="ID de grupos de Host")
parser.add_argument(
    "-i",
    "--data-inicio",
    help="Especifica uma data para inicio do periodo de extracao no formato dd/mm/YYYY [HH:MM:SS]",
)
parser.add_argument(
    "-f",
    "--data-fim",
    help="Data para o fim do periodo de extracao dd/mm/YYYY [HH:MM:SS]",
)
parser.add_argument(
    "-a",
    "--ack",
    action="store_true",
    help="Retorna apenas alertas que possuem acknowledged",
)
parser.add_argument("-e", "--email", help="Email para envio do relatório")
parser.add_argument("-l", "--last", type=int, help="Ultimas horas de alertas")
args = vars(parser.parse_args())

f = Figlet(font="slant")
print(f.renderText("Zabbix Export Events"))

if args["last"]:
    if args["data_inicio"] or args["data_fim"]:
        print("Conflito de parametos! Use somente --last ou --data-inicio/fim")
        exit()
    else:
        now = datetime.now()
        data_inicio = now - timedelta(hours=args["last"])
        data_final = now
        DATA_INICIO = time.mktime(
            time.strptime(
                data_inicio.strftime("%d/%m/%Y %H:%M:%S"), "%d/%m/%Y %H:%M:%S"
            )
        )
        DATA_FIM = time.mktime(
            time.strptime(data_final.strftime("%d/%m/%Y %H:%M:%S"), "%d/%m/%Y %H:%M:%S")
        )

# CONECTANDO AO ZABBIX
zapi = ZabbixAPI(args["server"])
if "https" in args["server"]:
    import requests

    requests.packages.urllib3.disable_warnings()
    zapi.session.verify = False
zapi.login(args["user"], args["password"])
print("--> Conectado com sucesso!\n")


def get_group_ids(group="*"):
    ids = []
    if group == "*":
        for g in zapi.hostgroup.get(
            output=["name"],
            search={"name": group},
            searchByAny=True,
            searchWildcardsEnabled=True,
        ):
            if "Template" not in g["name"]:
                ids.append(g["groupid"])
    else:
        for g in zapi.hostgroup.get(
            output=["name"], search={"name": group}, searchByAny=True
        ):
            if "Template" not in g["name"]:
                ids.append(g["groupid"])
    return ids


if args["group"]:
    grupos = get_group_ids(args["group"])
else:
    grupos = get_group_ids()

for g in zapi.hostgroup.get(output="extend", groupids=grupos):
    print("--> Grupo selecionado: " + g["name"])

# CRIANDO PLANILHA
wb = Workbook()
sheet = wb.active
sheet.title = "Eventos"
sheet["A1"] = "INICIO"
sheet["A1"].font = Font(sz=12, bold=True)
sheet["B1"] = "FIM"
sheet["B1"].font = Font(sz=12, bold=True)
sheet["C1"] = "HOST"
sheet["C1"].font = Font(sz=12, bold=True)
sheet["D1"] = "TRIGGER"
sheet["D1"].font = Font(sz=12, bold=True)
sheet["E1"] = "SEVERIDADE"
sheet["E1"].font = Font(sz=12, bold=True)
sheet["F1"] = "STATUS"
sheet["F1"].font = Font(sz=12, bold=True)
sheet["G1"] = "GRUPO"
sheet["G1"].font = Font(sz=12, bold=True)
sheet["H1"] = "APP"
sheet["H1"].font = Font(sz=12, bold=True)
sheet["I1"] = "ACK"
sheet["I1"].font = Font(sz=12, bold=True)
sheet["J1"] = "DATA NOTA"
sheet["J1"].font = Font(sz=12, bold=True)
sheet["K1"] = "NOTA"
sheet["K1"].font = Font(sz=12, bold=True)


# CONFIGURANDO DATA DE INICIO DO PERIODO
if args["data_inicio"]:
    try:
        args["data_inicio"] == datetime.strptime(
            args["data_inicio"], "%d/%m/%Y"
        ).strftime("%d/%m/%Y")
        DATA_INICIO = time.mktime(
            time.strptime(
                "{} 00:00:00".format(args["data_inicio"]), "%d/%m/%Y %H:%M:%S"
            )
        )
    except:
        args["data_inicio"] == datetime.strptime(
            args["data_inicio"], "%d/%m/%Y %H:%M:%S"
        ).strftime("%d/%m/%Y %H:%M:%S")
        DATA_INICIO = time.mktime(
            time.strptime(args["data_inicio"], "%d/%m/%Y %H:%M:%S")
        )


# CONFIGURANDO DATA FINAL DO PERIODO
if args["data_fim"]:
    try:
        args["data_fim"] == datetime.strptime(args["data_fim"], "%d/%m/%Y").strftime(
            "%d/%m/%Y"
        )
        DATA_FIM = time.mktime(
            time.strptime("{} 00:00:00".format(args["data_fim"]), "%d/%m/%Y %H:%M:%S")
        )
    except:
        args["data_fim"] == datetime.strptime(
            args["data_fim"], "%d/%m/%Y %H:%M:%S"
        ).strftime("%d/%m/%Y %H:%M:%S")
        DATA_FIM = time.mktime(time.strptime(args["data_fim"], "%d/%m/%Y %H:%M:%S"))


# VALOR EVENTO
EVENT = {"0": "OK", "1": "PROBLEM"}

# SEVERIDADE
SEV = {
    "0": "Not classified",
    "1": "Information",
    "2": "Warning",
    "3": "Average",
    "4": "High",
    "5": "Disaster",
}

# ACKNOWLEDGED
ACK = {"0": "Não", "1": "Sim"}

row = 2
eventos = 0
Col_Host = 0
Col_Trigger = 0
Col_Grupo = 0
Col_App = 0
Col_Message = 0

max = len(
    zapi.event.get(
        output=["eventid"],
        time_from=DATA_INICIO,
        time_till=DATA_FIM,
        value=1,
        groupids=grupos,
        acknowledged=args["ack"],
    )
)

print("\n")
bar = FillingSquaresBar(
    "--> Gerado relatorio de eventos!", max=max, suffix="%(percent).1f%% - %(elapsed)ds"
)


for e in zapi.event.get(
    output="extend",
    time_from=DATA_INICIO,
    time_till=DATA_FIM,
    sortfield=["clock"],
    sortorder="ASC",
    value=1,
    groupids=grupos,
    select_acknowledges="extend",
    acknowledged=args["ack"],
):
    for t in zapi.trigger.get(
        output="extend",
        triggerids=e["objectid"],
        expandDescription=True,
        min_severity=4,
        selectFunctions="extend",
    ):
        for h in zapi.host.get(
            output="extend", triggerids=e["objectid"], selectGroups=["name"]
        ):
            data = time.strftime("%d/%m/%Y %H:%M:%S", time.localtime(int(e["clock"])))
            fim = zapi.event.get(output=["clock"], eventids=e["r_eventid"])
            if len(fim) > 0:
                data_fim = time.strftime(
                    "%d/%m/%Y %H:%M:%S", time.localtime(int(fim[0]["clock"]))
                )
            else:
                data_fim = "N/A"
            app = zapi.application.get(
                output=["name"], itemids=t["functions"][0]["itemid"], limit=1
            )
            if len(app) > 0:
                sheet.cell(row=row, column=1).value = data
                sheet.cell(row=row, column=2).value = data_fim
                sheet.cell(row=row, column=3).value = h["host"]
                if len(h["host"]) > Col_Host:
                    Col_Host = len(h["host"])
                sheet.cell(row=row, column=4).value = t["description"]
                if len(t["description"]) > Col_Trigger:
                    Col_Trigger = len(t["description"])
                sheet.cell(row=row, column=5).value = SEV[t["priority"]]
                sheet.cell(row=row, column=6).value = EVENT[e["value"]]
                sheet.cell(row=row, column=7).value = h["groups"][0]["name"]
                if len(h["groups"][0]["name"]) > Col_Grupo:
                    Col_Grupo = len(h["groups"][0]["name"])
                sheet.cell(row=row, column=8).value = app[0]["name"]
                if len(app[0]["name"]) > Col_App:
                    Col_App = len(app[0]["name"])
                sheet.cell(row=row, column=9).value = ACK[e["acknowledged"]]
                if len(e["acknowledges"]) > 0:
                    sheet.cell(row=row, column=10).value = time.strftime(
                        "%d/%m/%Y %H:%M:%S",
                        time.localtime(int(e["acknowledges"][0]["clock"])),
                    )
                    sheet.cell(row=row, column=11).value = e["acknowledges"][0][
                        "message"
                    ]
                    if len(e["acknowledges"][0]["message"]) > Col_Message:
                        Col_Message = len(e["acknowledges"][0]["message"])
            else:
                sheet.cell(row=row, column=1).value = data
                sheet.cell(row=row, column=2).value = data_fim
                sheet.cell(row=row, column=3).value = h["host"]
                if len(h["host"]) > Col_Host:
                    Col_Host = len(h["host"])
                sheet.cell(row=row, column=4).value = t["description"]
                if len(t["description"]) > Col_Trigger:
                    Col_Trigger = len(t["description"])
                sheet.cell(row=row, column=5).value = SEV[t["priority"]]
                sheet.cell(row=row, column=6).value = EVENT[e["value"]]
                sheet.cell(row=row, column=7).value = h["groups"][0]["name"]
                if len(h["groups"][0]["name"]) > Col_Grupo:
                    Col_Grupo = len(h["groups"][0]["name"])
                sheet.cell(row=row, column=8).value = "N/A"
                if len(h["groups"][0]["name"]) > Col_Grupo:
                    Col_Grupo = len(h["groups"][0]["name"])
                sheet.cell(row=row, column=9).value = ACK[e["acknowledged"]]
                if len(e["acknowledges"]) > 0:
                    sheet.cell(row=row, column=10).value = time.strftime(
                        "%d/%m/%Y %H:%M:%S",
                        time.localtime(int(e["acknowledges"][0]["clock"])),
                    )
                    sheet.cell(row=row, column=11).value = e["acknowledges"][0][
                        "message"
                    ]
                    if len(e["acknowledges"][0]["message"]) > Col_Message:
                        Col_Message = len(e["acknowledges"][0]["message"])

            row += 1
            eventos += 1
            bar.next()

bar.finish()
area = sheet.dimensions
sheet.auto_filter.ref = area
sheet.column_dimensions["A"].width = 18
sheet.column_dimensions["B"].width = 18
sheet.column_dimensions["C"].width = Col_Host + 3
sheet.column_dimensions["D"].width = Col_Trigger
sheet.column_dimensions["E"].width = 15
sheet.column_dimensions["F"].width = 15
sheet.column_dimensions["G"].width = Col_Grupo + 3
sheet.column_dimensions["H"].width = Col_App + 1
sheet.column_dimensions["I"].width = 7
sheet.column_dimensions["J"].width = 18
sheet.column_dimensions["K"].width = Col_Message
sheet.freeze_panes = "A2"


DATE = date.today()
if args["name"]:
    NOME = "Alertas_" + args["name"] + "_" + str(DATE) + ".xlsx"
else:
    NOME = "Alertas_" + str(DATE) + ".xlsx"

wb.save(NOME)
zapi.user.logout()


# ENVIO DO RELATORIO POR Email
if args["email"]:
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    import smtplib

    email_user = "account@company.com"  # Account used for send e-mail
    password = "password"  # Account passowrd used to send e-mail
    smtp_adrress = "smtp.company.com"  # SMTP Server
    email_send = args["email"]  # E-mail to receve e-mail

    msg = MIMEMultipart()

    message = "Relatorio {} gerado! Segue em anexo".format(NOME)

    msg["From"] = "Account Name <{}>".format(email_user)
    msg["To"] = email_send
    msg["Subject"] = "Relatorio de alertas"

    msg.attach(MIMEText(message, "plain"))

    anexo = open(NOME, "rb")

    part = MIMEBase("application", "octet-stream")
    part.set_payload((anexo).read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment; filename=" + NOME)

    msg.attach(part)

    server = smtplib.SMTP(smtp_adrress, 587)
    server.starttls()
    server.login(email_user, password)

    server.sendmail(msg["From"], msg["To"], msg.as_string())

    server.quit()

    print("\n--> Relatorio {} gerado e enviado com sucesso!".format(NOME))

else:
    print("\n--> Relatorio {} gerado com sucesso!".format(NOME))
